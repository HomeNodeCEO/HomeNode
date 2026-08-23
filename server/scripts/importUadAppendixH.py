#!/usr/bin/env python3
"""Import the official joint-GSE Appendix H-1 workbook without executing VBA.

This is a maintainer tool, not a runtime dependency. It writes a deterministic
JSON source catalog and additive PostgreSQL migration. Run it only against a
workbook downloaded from one of the pinned official GSE URLs.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


RELEASE_KEY = "uad-3.6-2026-08-13-h1.5"
APPENDIX_VERSION = "1.5"
PUBLISHED_ON = "2026-08-13"
EXPECTED_RULE_COUNT = 728
EXPECTED_FATAL_COUNT = 592
EXPECTED_WARNING_COUNT = 136
SHEET_NAME = "UAD Compliance Rules v1.5"
FANNIE_URL = "https://singlefamily.fanniemae.com/media/document/xlsm/appendix-h1-uad-compliance-rules-urar"
FREDDIE_URL = "https://sf.freddiemac.com/docs/xlsx/appendix-h-1-uad-compliance-rules-urar.xlsx"
DELETED_RULE_IDS = ("UAD1438", "UAD1443", "UAD1625")
TARGET_MIGRATION = "20260926_uad_appendix_h1_v1_5.sql"


def normalized(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").strip()
    return value


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def canonical_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def existing_local_rule_ids(migrations: Path) -> set[str]:
    rule_ids: set[str] = set()
    for migration in sorted(migrations.glob("*.sql")):
        if migration.name >= TARGET_MIGRATION:
            continue
        contents = migration.read_text(encoding="utf-8")
        rule_ids.update(re.findall(r"['\"](UAD\d+)['\"]", contents))
    return rule_ids


def load_rules(source: Path, local_rule_ids: set[str]):
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"missing official sheet: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required_headers = {
        "Unique ID", "Primary Data Element", "Message ID", "Message Text",
        "Rule Logic", "Severity", "Property Affected", "Report Section",
        "Report Subsection", "Report Label / Value", "Data Point Name / Value",
        "xPath", "Related Value(s)", "Date Format", "Number Format",
        "Min Value", "Max Value",
    }
    if set(headers) != required_headers:
        raise SystemExit("Appendix H-1 column contract changed; review the importer before continuing")

    rules = []
    for source_row, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: normalized(cells[index]) for index, header in enumerate(headers)}
        rule_id = str(row["Message ID"] or "")
        if not re.fullmatch(r"UAD\d{4}", rule_id):
            raise SystemExit(f"unexpected rule ID at row {source_row}: {rule_id!r}")
        severity = str(row["Severity"] or "").lower()
        if severity not in {"fatal", "warning"}:
            raise SystemExit(f"unexpected severity for {rule_id}: {severity!r}")
        source_values = [row[header] for header in headers]
        fingerprint = sha256_bytes(canonical_json(source_values).encode("utf-8"))
        rules.append({
            "rule_id": rule_id,
            "unique_id": str(row["Unique ID"] or ""),
            "primary_data_element": row["Primary Data Element"],
            "severity": severity,
            "property_affected": row["Property Affected"],
            "message": row["Message Text"],
            "rule_logic": row["Rule Logic"],
            "report_section": row["Report Section"],
            "report_subsection": row["Report Subsection"],
            "report_label": row["Report Label / Value"],
            "data_point": row["Data Point Name / Value"],
            "xpath": row["xPath"],
            "related_values": row["Related Value(s)"],
            "date_format": row["Date Format"],
            "number_format": row["Number Format"],
            "min_value": row["Min Value"],
            "max_value": row["Max Value"],
            "source_row": source_row,
            "source_fingerprint_sha256": fingerprint,
            "local_evaluation_status": (
                "mapped_unverified" if rule_id in local_rule_ids else "reference_only"
            ),
        })

    rule_ids = [rule["rule_id"] for rule in rules]
    severities = Counter(rule["severity"] for rule in rules)
    if len(rules) != EXPECTED_RULE_COUNT or len(set(rule_ids)) != EXPECTED_RULE_COUNT:
        raise SystemExit(f"expected {EXPECTED_RULE_COUNT} unique rules, found {len(set(rule_ids))}")
    if severities != Counter({"fatal": EXPECTED_FATAL_COUNT, "warning": EXPECTED_WARNING_COUNT}):
        raise SystemExit(f"unexpected severity totals: {dict(severities)}")
    deleted_present = sorted(set(rule_ids).intersection(DELETED_RULE_IDS))
    if deleted_present:
        raise SystemExit(f"deleted rules unexpectedly present: {', '.join(deleted_present)}")
    return rules


def source_row(rule):
    metadata = {
        "source": f"Appendix H-1 v{APPENDIX_VERSION}",
        "source_row": rule["source_row"],
        "primary_data_element": rule["primary_data_element"],
        "report_section": rule["report_section"],
        "report_subsection": rule["report_subsection"],
        "report_label": rule["report_label"],
        "data_point": rule["data_point"],
        "xpath": rule["xpath"],
        "related_values": rule["related_values"],
        "date_format": rule["date_format"],
        "number_format": rule["number_format"],
        "min_value": rule["min_value"],
        "max_value": rule["max_value"],
    }
    return {
        "rule_id": rule["rule_id"],
        "severity": rule["severity"],
        "property_context": rule["property_affected"],
        "message": rule["message"],
        "expression": rule["rule_logic"],
        "source_unique_id": rule["unique_id"],
        "source_fingerprint_sha256": rule["source_fingerprint_sha256"],
        "local_evaluation_status": rule["local_evaluation_status"],
        "metadata": metadata,
    }


def migration_sql(catalog, rules_json: str) -> str:
    document = catalog["document"]
    deleted = ", ".join(f"'{rule_id}'" for rule_id in DELETED_RULE_IDS)
    return f"""-- Official Appendix H-1 v{APPENDIX_VERSION} source catalog for the URAR.
-- This imports reference data only. `mapped_unverified` records identify rules
-- with an existing HomeNode mapping; they are not claimed as exact GSE-equivalent
-- execution until promoted by rule-specific conformance evidence.

ALTER TABLE uad_ref.compliance_rules
  ADD COLUMN IF NOT EXISTS source_unique_id text,
  ADD COLUMN IF NOT EXISTS source_fingerprint_sha256 text,
  ADD COLUMN IF NOT EXISTS local_evaluation_status text NOT NULL DEFAULT 'reference_only';

ALTER TABLE uad_ref.compliance_rules
  DROP CONSTRAINT IF EXISTS uad_compliance_rules_source_fingerprint_check;
ALTER TABLE uad_ref.compliance_rules
  ADD CONSTRAINT uad_compliance_rules_source_fingerprint_check
  CHECK (source_fingerprint_sha256 IS NULL OR source_fingerprint_sha256 ~ '^[0-9a-f]{{64}}$');

ALTER TABLE uad_ref.compliance_rules
  DROP CONSTRAINT IF EXISTS uad_compliance_rules_local_evaluation_status_check;
ALTER TABLE uad_ref.compliance_rules
  ADD CONSTRAINT uad_compliance_rules_local_evaluation_status_check
  CHECK (local_evaluation_status IN ('reference_only', 'mapped_unverified', 'locally_verified'));

CREATE TABLE IF NOT EXISTS uad_ref.compliance_rule_source_manifests (
  release_key text NOT NULL
    REFERENCES uad_ref.specification_releases(release_key) ON DELETE CASCADE,
  appendix_key text NOT NULL,
  document_version text NOT NULL,
  published_on date NOT NULL,
  official_source_url text NOT NULL,
  alternate_official_source_url text,
  source_file_name text NOT NULL,
  source_sha256 text NOT NULL,
  rule_catalog_sha256 text NOT NULL,
  active_rule_count integer NOT NULL,
  fatal_rule_count integer NOT NULL,
  warning_rule_count integer NOT NULL,
  imported_at timestamptz NOT NULL DEFAULT now(),
  PRIMARY KEY (release_key, appendix_key),
  CHECK (source_sha256 ~ '^[0-9a-f]{{64}}$'),
  CHECK (rule_catalog_sha256 ~ '^[0-9a-f]{{64}}$'),
  CHECK (active_rule_count > 0),
  CHECK (fatal_rule_count >= 0 AND warning_rule_count >= 0),
  CHECK (fatal_rule_count + warning_rule_count = active_rule_count)
);

WITH official AS (
  SELECT * FROM jsonb_to_recordset($appendix_h_1${rules_json}$appendix_h_1$::jsonb) AS rule(
    rule_id text,
    severity text,
    property_context text,
    message text,
    expression text,
    source_unique_id text,
    source_fingerprint_sha256 text,
    local_evaluation_status text,
    metadata jsonb
  )
)
INSERT INTO uad_ref.compliance_rules (
  release_key, rule_id, severity, property_context, message, expression,
  report_field_ids, metadata, source_unique_id, source_fingerprint_sha256,
  local_evaluation_status
)
SELECT
  '{RELEASE_KEY}', rule_id, severity, property_context, message, expression,
  ARRAY[]::text[], metadata, source_unique_id, source_fingerprint_sha256,
  local_evaluation_status
FROM official
ON CONFLICT (release_key, rule_id) DO UPDATE
SET severity = EXCLUDED.severity,
    property_context = EXCLUDED.property_context,
    message = EXCLUDED.message,
    expression = EXCLUDED.expression,
    metadata = uad_ref.compliance_rules.metadata || EXCLUDED.metadata,
    source_unique_id = EXCLUDED.source_unique_id,
    source_fingerprint_sha256 = EXCLUDED.source_fingerprint_sha256,
    local_evaluation_status = EXCLUDED.local_evaluation_status;

DELETE FROM uad_ref.compliance_rules
 WHERE release_key = '{RELEASE_KEY}'
   AND rule_id IN ({deleted});

INSERT INTO uad_ref.compliance_rule_source_manifests (
  release_key, appendix_key, document_version, published_on,
  official_source_url, alternate_official_source_url, source_file_name,
  source_sha256, rule_catalog_sha256, active_rule_count,
  fatal_rule_count, warning_rule_count
) VALUES (
  '{RELEASE_KEY}', 'H-1', '{APPENDIX_VERSION}', DATE '{PUBLISHED_ON}',
  '{FREDDIE_URL}', '{FANNIE_URL}', '{document["source_file_name"]}',
  '{document["source_sha256"]}', '{document["rule_catalog_sha256"]}',
  {EXPECTED_RULE_COUNT}, {EXPECTED_FATAL_COUNT}, {EXPECTED_WARNING_COUNT}
)
ON CONFLICT (release_key, appendix_key) DO UPDATE
SET document_version = EXCLUDED.document_version,
    published_on = EXCLUDED.published_on,
    official_source_url = EXCLUDED.official_source_url,
    alternate_official_source_url = EXCLUDED.alternate_official_source_url,
    source_file_name = EXCLUDED.source_file_name,
    source_sha256 = EXCLUDED.source_sha256,
    rule_catalog_sha256 = EXCLUDED.rule_catalog_sha256,
    active_rule_count = EXCLUDED.active_rule_count,
    fatal_rule_count = EXCLUDED.fatal_rule_count,
    warning_rule_count = EXCLUDED.warning_rule_count,
    imported_at = now();

UPDATE uad_ref.specification_releases
   SET source_manifest = source_manifest || jsonb_build_object(
         'appendix_h_1', jsonb_build_object(
           'version', '{APPENDIX_VERSION}',
           'published_on', '{PUBLISHED_ON}',
           'active_rule_count', {EXPECTED_RULE_COUNT},
           'fatal_rule_count', {EXPECTED_FATAL_COUNT},
           'warning_rule_count', {EXPECTED_WARNING_COUNT},
           'source_sha256', '{document["source_sha256"]}',
           'rule_catalog_sha256', '{document["rule_catalog_sha256"]}'
         )
       ),
       source_manifest_sha256 = '{document["rule_catalog_sha256"]}',
       imported_at = now()
 WHERE release_key = '{RELEASE_KEY}';
"""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument(
        "--catalog",
        type=Path,
        default=Path("src/modules/uad/spec/appendix-h1-v1.5.json"),
    )
    parser.add_argument(
        "--migration",
        type=Path,
        default=Path(f"migrations/{TARGET_MIGRATION}"),
    )
    args = parser.parse_args()
    if not args.source.is_file():
        raise SystemExit(f"source workbook not found: {args.source}")

    migrations = args.migration.parent.resolve()
    local_rule_ids = existing_local_rule_ids(migrations)
    rules = load_rules(args.source, local_rule_ids)
    rules_hash = sha256_bytes(canonical_json(rules).encode("utf-8"))
    mapped = sum(rule["local_evaluation_status"] == "mapped_unverified" for rule in rules)
    catalog = {
        "schema_version": 1,
        "document": {
            "appendix": "H-1",
            "title": "UAD Compliance Rules - Uniform Residential Appraisal Report",
            "version": APPENDIX_VERSION,
            "published_on": PUBLISHED_ON,
            "release_key": RELEASE_KEY,
            "official_source_url": FREDDIE_URL,
            "alternate_official_source_url": FANNIE_URL,
            "source_file_name": args.source.name,
            "source_sha256": sha256_bytes(args.source.read_bytes()),
            "rule_catalog_sha256": rules_hash,
            "active_rule_count": EXPECTED_RULE_COUNT,
            "fatal_rule_count": EXPECTED_FATAL_COUNT,
            "warning_rule_count": EXPECTED_WARNING_COUNT,
            "deleted_rule_ids": list(DELETED_RULE_IDS),
        },
        "coverage": {
            "cataloged_rule_count": EXPECTED_RULE_COUNT,
            "mapped_unverified_rule_count": mapped,
            "reference_only_rule_count": EXPECTED_RULE_COUNT - mapped,
            "locally_verified_rule_count": 0,
            "gse_equivalence_claimed": False,
        },
        "rules": rules,
    }
    args.catalog.parent.mkdir(parents=True, exist_ok=True)
    args.migration.parent.mkdir(parents=True, exist_ok=True)
    args.catalog.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    sql_rows = [source_row(rule) for rule in rules]
    args.migration.write_text(
        migration_sql(catalog, canonical_json(sql_rows)),
        encoding="utf-8",
        newline="\n",
    )
    print(json.dumps({
        "ok": True,
        "source_sha256": catalog["document"]["source_sha256"],
        "rule_catalog_sha256": rules_hash,
        "active_rule_count": EXPECTED_RULE_COUNT,
        "mapped_unverified_rule_count": mapped,
        "reference_only_rule_count": EXPECTED_RULE_COUNT - mapped,
        "catalog": str(args.catalog),
        "migration": str(args.migration),
    }, indent=2))


if __name__ == "__main__":
    main()
